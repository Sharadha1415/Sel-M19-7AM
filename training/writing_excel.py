'''
openpyxl    :   To write into the excel files, we use openpyxl
                *   Workbook()  --> creates a new excel workbook
                *   .active     --> gets the current sheet

                To install openpyxl
                Go to command prompt    -->     pip install openpyxl

'''

from openpyxl import Workbook

## create a new excel workbook
workbook = Workbook()

## initialize the worksheet
worksheet = workbook.active

## set the title for the worksheet(optional)
worksheet.title = "personal_info"

## write the data into the excel
worksheet['A1'] = 'name'
worksheet['B1'] = 'place'
worksheet['C1'] = 'email_id'
worksheet['D1'] = 'phone_number'

data_list = [
    ['Venkat', 'Hyderabad', 'venkat@gmail.com', 9874102364],
    ['Nagesh', 'Mumbai', 'nagesh@gmail.com', 7896302541],
    ['Yamini', 'Chennai', 'yamini@gmail.com', 8520147963],
    ['Shaheed', 'Delhi', 'shaheed@gmail.com', 9635874120],
    ['Sumeet', 'Bengaluru', 'sumeet@gmail.com', 4569871230],
    ['Shravya', 'Kolkatta', 'shravya@gmail.com', 7469852310],
    ['Mainak', 'Pune', 'mainak@gmail.com', 8976542310],
    ['Chandrashekar', 'Mysuru', 'chandrashekar@gmail.com', 7698452103]
]

for data in data_list:
    worksheet.append(data)

# ## save the excel file
# workbook.save("M19.xlsx")

## To save the excel file in different location
workbook.save(r'C:\Users\Ramya\PycharmProjects\Sel-M19-Jan5-7AM\files\m19.xlsx')

############################################################################################

# ## ANALYZE THE BELOW CODE
#
# from openpyxl import Workbook
# from openpyxl.reader.excel import load_workbook
# from openpyxl.styles import Font, PatternFill, Border, Alignment
#
# ## create the excel workbook
# excel_workbook = Workbook()
#
# ## initialize the worksheet
# worksheet = excel_workbook.active
#
# ## set a name for the worksheet(optional)
# worksheet.title = 'personal_information'
#
# ## Adding the data
# worksheet['A1'] = 'name'
# worksheet['B1'] = 'place'
# worksheet['C1'] = 'salary'
# worksheet['D1'] = 'email_id'
#
# ## Font formatting
# worksheet['A1'].font = Font(name='Arial', bold=True, italic=True, color='FF0000')   ## formats individual cell
# worksheet['B1'].font = Font(name='Forte', bold=True, italic=True, color='F54927')   ## formats individual cell
#
#
# data_list = [
#     ['Vihaan', 'Delhi', 50000, 'vihaan@gmail.com'],
#     ['Siddharth', 'Mumbai', 55000, 'siddu@gmail.com'],
#     ['Saina', 'Hyderabad', 60000, 'saina@gmail.com'],
#     ['Risto', 'Pune', 65000, 'risto@gmail.com'],
#     ['Goku', 'Assam', 70000, 'goku@gmail.com'],
#     ['Paresh', 'Gujrat', 75000, 'paresh@gmail.com']
# ]
#
# for row in data_list:
#     worksheet.append(row)
#
# excel_workbook.save('file1.xlsx')

#-----------------------------------------------------------------------------
# ## formatting entire row
# ## define styles
# bold_font = Font(bold=True, color='FFFFFF')
# fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
# for cell in worksheet[1]:
#     cell.font = bold_font
#     cell.fill = fill

# #-----------------------------------------------------------------------------

# ## Add new sheet to the existing workbook
#
# ## load existing workbook
# workbook = load_workbook('emp_data.xlsx')
# # workbook = load_workbook('path\emp_data.xlsx')
#
# ## create a new sheet
# new_sheet = workbook.create_sheet(title="title")
#
# ## write the data asusal
# ## save the file

# ##-----------------------------------------------------------------------------
#
# ## save the excel file
# excel_workbook.save('emp_data.xlsx')
#
# # ## save the excel file in different location
# # excel_workbook.save(r'C:\Users\Ramya\PycharmProjects\Sel-M6-April1\files_\emp_data.xlsx')
#
# ##-----------------------------------------------------------------------------

# ## Clear the cell's value
#
# from openpyxl import load_workbook
#
# workbook = load_workbook("emp_data.xlsx")
# worksheet = workbook.active
#
# worksheet['A1'] = None      ## Clears the value and formats in cell A1
#
# workbook.save("emp_data.xlsx")

# ##-----------------------------------------------------------------------------

# ## Delete entire row or column
#
# from openpyxl import load_workbook
#
# workbook = load_workbook("emp_data.xlsx")
# worksheet = workbook.active
#
# worksheet.delete_rows(1)        ## Deletes row1
# worksheet.delete_cols(1)        ## Deletes col1























































